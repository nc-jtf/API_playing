from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side
from scripts.database_connection import *
import pandas as pd
import numpy as np
from sklearn.datasets import load_iris
import sklearn

# def find_value_in_excel(file_path, sheet_name, target_value):
#     # Open the Excel file
#     workbook = load_workbook(file_path)
#
#     # Select the sheet
#     sheet = workbook[sheet_name]
#     rows = sheet.iter_rows()
#
#     for row in sheet.iter_rows():
#         print(f"{sheet.cell.value}")
#         # return sheet.cell.value
def read_using_pd(file_path):
    df = pd.read_excel(file_path, header=2)
    # print(df.columns)
    thisdict = {}
    filtered_data = df[df['CVR-nr. certifikat'].isin(cvr_list)]
    for key,row in filtered_data.iterrows():
        firmanavn = row['Firmanavn']
        cvr = str(row['CVR-nr. certifikat'])
        thisdict.update({cvr: firmanavn})
    return thisdict
    #     value_at_row_column = df.at[2, 'Virksomheder til onboarding']  # Unnamed: 56
    #     print(value_at_row_column)
    #     print(df.head(5))
# def extract_cvr_company_names(file_path, cvr_list):
#     # Read the Excel file into a DataFrame
#     dataframe = pd.read_excel(file_path, header=2)
#
#     # Ensure the headers are in the expected format
#     # if 'CVR-nr. certifikat' not in dataframe.columns or 'Firmanavn' not in dataframe.columns:
#     #     raise ValueError("Expected headers not found in the Excel file.")
#
#     # Filter the DataFrame based on the provided CVR list
#     filtered_data = dataframe[dataframe['CVR-nr. certifikat'].isin(cvr_list)]
#
#     # Create a dictionary with CVR numbers as keys and corresponding company names as values
#     cvr_company_dict = dict(zip(filtered_data['CVR-nr. certifikat'], filtered_data['Firmanavn']))
#
#     return cvr_company_dict
if __name__ == '__main__':
    cvr_list = ['10701732', '83031212', '78933917', '16993409']
    file_path = fr"C:\Users\jtf\Downloads\EO access status overview.xlsx"
    sheet_name = "Virksomhed"
    dict = read_using_pd(file_path)
    # dict = extract_cvr_company_names(file_path, cvr_list)