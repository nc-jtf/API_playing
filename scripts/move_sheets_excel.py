from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, Border, Side
from scripts.database_connection import *
def pack_out(list) -> list:
    '''Pakker en liste af tuples[0] med string ud til en liste med strings'''
    new_list = []
    for tup in list:
        new_list.append(tup[0])
    return new_list

def create_excel(domain=None):
    cur = cursor("tfe01_ref_data_mgt_bs")
    # 1. find unikke codelist_keys
    if domain == "EX" or domain == "export":
        domain_val = "AES-"
        domain_name = " - AES"
    if domain == "TRA" or domain == "transit":
        domain_val = "NCTS-P5-"
        domain_name = " - NCTS"
    if domain == "GLOBAL":
        domain_val = "GLOBAL-"
        domain_name = " - Global"
    if domain == None:
        domain_val = ""
        domain_name = ""

    unique_sql_all = f'''SELECT DISTINCT codelist.codelist_key
    FROM codelist_item
    JOIN codelist_item_label ON codelist_item.id = codelist_item_label.item_id
    JOIN codelist ON codelist_item.codelist_id = codelist.id
    JOIN codelist_label ON codelist_label.codelist_id = codelist.id
    WHERE codelist_item_label.language_id = 1
      AND codelist.codelist_key LIKE '%{domain_val}CL%'
      AND codelist.codelist_key NOT LIKE '%DEPRECATED%'
    ORDER BY codelist.codelist_key ASC'''

    cur.execute(unique_sql_all)
    unique_keys = cur.fetchall()
    unique_keys = pack_out(unique_keys)
    i = 1
    # 2. For hver unik key, lav en sheet med det navn
    workbook = Workbook()

    # Remove the default sheet created by openpyxl
    workbook.remove(workbook.active)

    # Create a sheet for each name in the list
    for sheet_name in unique_keys:
        print(sheet_name)
        workbook.create_sheet(title=sheet_name)
        sheet = workbook[sheet_name]
        # Borders
        border = Border(
            left=Side(border_style="thin"),
            right=Side(border_style="thin"),
            top=Side(border_style="thin"),
            bottom=Side(border_style="thin")
        )

        start_cell = 'A1'
        end_cell = 'B1'
        sheet.merge_cells(f'{start_cell}:{end_cell}')
        sheet.cell(row=1, column=1, value=sheet_name)
        sheet.cell(row=1, column=1).font = Font(size=16, bold=True)

        # "Code" cell
        sheet.cell(row=2, column=1, value="Code")
        sheet.cell(row=2, column=1).font = Font(bold=True)
        sheet.cell(row=2, column=1).border = border

        # "Description" cell
        sheet.cell(row=2, column=2, value="Description")
        sheet.cell(row=2, column=2).font = Font(bold=True)
        sheet.cell(row=2, column=2).border = border

        sql_specific_cl = f'''SELECT codelist_item.item_code, codelist_item_label.description
        FROM codelist_item
        JOIN codelist_item_label ON codelist_item.id = codelist_item_label.item_id
        JOIN codelist ON codelist_item.codelist_id = codelist.id
        JOIN codelist_label ON codelist_label.codelist_id = codelist.id
        WHERE codelist_item_label.language_id = 1
          AND codelist.codelist_key LIKE '%{sheet_name}%'
        ORDER BY codelist.codelist_key ASC'''
        cur = cursor("tfe01_ref_data_mgt_bs")
        cur.execute(sql_specific_cl)
        values = cur.fetchall()
        start_row = 3

        for item_code, description in values:
            sheet.cell(row=start_row, column=1, value=item_code)
            sheet.cell(row=start_row, column=2, value=description)
            sheet.cell(row=start_row, column=1).border = border
            sheet.cell(row=start_row, column=2).border = border
            start_row += 1


    # Save the workbook
    workbook.save(fr"C:\Users\jtf\Downloads\Codelists{domain_name}.xlsx")
    print("Workbook saved")

if __name__ == '__main__':
    create_excel("EX")
    create_excel("TRA")
    create_excel("GLOBAL")